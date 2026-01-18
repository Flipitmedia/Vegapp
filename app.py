"""
Sistema de Gestión de Pedidos - La Vega
Procesa exportaciones de Shopify y genera listas de compras y armado.
"""

from fastapi import FastAPI, UploadFile, File, Request, Form, HTTPException
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
import csv
import io
import re
from datetime import datetime, date
from typing import Optional
import json
import os
from pathlib import Path

# Para generar Excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Base de datos SQLite
import sqlite3

app = FastAPI(title="Sistema Gestión La Vega")

# Configurar archivos estáticos y templates
BASE_DIR = Path(__file__).resolve().parent
app.mount("/static", StaticFiles(directory=BASE_DIR / "static"), name="static")
templates = Jinja2Templates(directory=BASE_DIR / "templates")

# Directorio para archivos generados
OUTPUT_DIR = BASE_DIR / "outputs"
OUTPUT_DIR.mkdir(exist_ok=True)

DB_PATH = BASE_DIR / "vega.db"

# ============================================
# BASE DE DATOS
# ============================================

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    """Inicializa la base de datos con las tablas necesarias."""
    conn = get_db()
    cursor = conn.cursor()
    
    # Tabla de categorías
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS categorias (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT UNIQUE NOT NULL,
            orden INTEGER DEFAULT 0
        )
    ''')
    
    # Tabla de mapeo producto -> categoría
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS producto_categoria (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            producto TEXT UNIQUE NOT NULL,
            categoria_id INTEGER,
            FOREIGN KEY (categoria_id) REFERENCES categorias(id)
        )
    ''')
    
    # Tabla de pedidos
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS pedidos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            order_number TEXT UNIQUE NOT NULL,
            email TEXT,
            comuna TEXT,
            fecha_entrega DATE,
            direccion TEXT,
            telefono TEXT,
            nombre_cliente TEXT,
            total REAL,
            created_at TIMESTAMP,
            imported_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            status TEXT DEFAULT 'pendiente'
        )
    ''')
    
    # Tabla de líneas de pedido
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS lineas_pedido (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            pedido_id INTEGER,
            producto TEXT NOT NULL,
            cantidad INTEGER NOT NULL,
            precio REAL,
            sku TEXT,
            FOREIGN KEY (pedido_id) REFERENCES pedidos(id)
        )
    ''')
    
    # Insertar categorías por defecto si no existen
    categorias_default = [
        ('Frutas', 1),
        ('Verduras', 2),
        ('Congelados', 3),
        ('Abarrotes', 4),
        ('Lácteos', 5),
        ('Carnes', 6),
        ('Otros', 99)
    ]
    
    for nombre, orden in categorias_default:
        cursor.execute('''
            INSERT OR IGNORE INTO categorias (nombre, orden) VALUES (?, ?)
        ''', (nombre, orden))
    
    conn.commit()
    conn.close()

# Inicializar DB al arrancar
init_db()

# ============================================
# UTILIDADES
# ============================================

def parse_note_attributes(note_attrs: str) -> dict:
    """Extrae comuna y fecha de entrega de los note attributes."""
    result = {'comuna': None, 'fecha_entrega': None}
    
    if not note_attrs:
        return result
    
    # Buscar comuna
    comuna_match = re.search(r'Comuna de Entrega:\s*([^\n]+)', note_attrs)
    if comuna_match:
        result['comuna'] = comuna_match.group(1).strip()
    
    # Buscar fecha
    fecha_match = re.search(r'Fecha de Entrega:\s*(\d{4}-\d{2}-\d{2})', note_attrs)
    if fecha_match:
        result['fecha_entrega'] = fecha_match.group(1)
    
    return result

def parse_shopify_csv(content: str) -> list:
    """Parsea el CSV de Shopify y agrupa por pedido."""
    reader = csv.DictReader(io.StringIO(content))
    
    orders = {}
    
    for row in reader:
        order_number = row.get('Name', '')
        if not order_number:
            continue
            
        if order_number not in orders:
            note_attrs = parse_note_attributes(row.get('Note Attributes', ''))
            
            # Parsear fecha de creación
            created_at = None
            if row.get('Created at'):
                try:
                    created_at = datetime.strptime(
                        row['Created at'].split(' -')[0].split(' +')[0], 
                        '%Y-%m-%d %H:%M:%S'
                    )
                except:
                    pass
            
            orders[order_number] = {
                'order_number': order_number,
                'email': row.get('Email', ''),
                'comuna': note_attrs['comuna'],
                'fecha_entrega': note_attrs['fecha_entrega'],
                'nombre_cliente': row.get('Shipping Name', '') or row.get('Billing Name', ''),
                'direccion': row.get('Shipping Address1', ''),
                'telefono': row.get('Phone', '') or row.get('Shipping Phone', ''),
                'total': float(row.get('Total', 0) or 0),
                'created_at': created_at,
                'items': []
            }
        
        # Agregar línea de producto
        if row.get('Lineitem name'):
            orders[order_number]['items'].append({
                'producto': row['Lineitem name'],
                'cantidad': int(row.get('Lineitem quantity', 1) or 1),
                'precio': float(row.get('Lineitem price', 0) or 0),
                'sku': row.get('Lineitem sku', '')
            })
    
    return list(orders.values())

# ============================================
# RUTAS API
# ============================================

@app.get("/", response_class=HTMLResponse)
async def home(request: Request):
    """Página principal."""
    conn = get_db()
    cursor = conn.cursor()
    
    # Obtener estadísticas
    cursor.execute("SELECT COUNT(*) FROM pedidos WHERE status = 'pendiente'")
    pedidos_pendientes = cursor.fetchone()[0]
    
    cursor.execute("SELECT COUNT(DISTINCT fecha_entrega) FROM pedidos WHERE status = 'pendiente'")
    fechas_pendientes = cursor.fetchone()[0]
    
    # Pedidos de hoy
    hoy = date.today().isoformat()
    cursor.execute("SELECT COUNT(*) FROM pedidos WHERE fecha_entrega = ? AND status = 'pendiente'", (hoy,))
    pedidos_hoy = cursor.fetchone()[0]
    
    # Productos sin categoría
    cursor.execute('''
        SELECT COUNT(DISTINCT lp.producto) 
        FROM lineas_pedido lp
        LEFT JOIN producto_categoria pc ON lp.producto = pc.producto
        WHERE pc.id IS NULL
    ''')
    sin_categoria = cursor.fetchone()[0]
    
    conn.close()
    
    return templates.TemplateResponse("index.html", {
        "request": request,
        "pedidos_pendientes": pedidos_pendientes,
        "fechas_pendientes": fechas_pendientes,
        "pedidos_hoy": pedidos_hoy,
        "sin_categoria": sin_categoria,
        "fecha_hoy": hoy
    })


@app.post("/upload")
async def upload_csv(file: UploadFile = File(...)):
    """Sube y procesa un CSV de Shopify."""
    if not file.filename.endswith('.csv'):
        raise HTTPException(400, "El archivo debe ser CSV")
    
    content = await file.read()
    content = content.decode('utf-8-sig')  # Manejar BOM de Excel
    
    orders = parse_shopify_csv(content)
    
    conn = get_db()
    cursor = conn.cursor()
    
    nuevos = 0
    duplicados = 0
    sin_fecha = 0
    
    for order in orders:
        # Verificar si ya existe
        cursor.execute("SELECT id FROM pedidos WHERE order_number = ?", (order['order_number'],))
        existing = cursor.fetchone()
        
        if existing:
            duplicados += 1
            continue
        
        if not order['fecha_entrega']:
            sin_fecha += 1
            continue
        
        # Insertar pedido
        cursor.execute('''
            INSERT INTO pedidos (order_number, email, comuna, fecha_entrega, direccion, telefono, nombre_cliente, total, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            order['order_number'],
            order['email'],
            order['comuna'],
            order['fecha_entrega'],
            order['direccion'],
            order['telefono'],
            order['nombre_cliente'],
            order['total'],
            order['created_at']
        ))
        
        pedido_id = cursor.lastrowid
        
        # Insertar líneas
        for item in order['items']:
            cursor.execute('''
                INSERT INTO lineas_pedido (pedido_id, producto, cantidad, precio, sku)
                VALUES (?, ?, ?, ?, ?)
            ''', (pedido_id, item['producto'], item['cantidad'], item['precio'], item['sku']))
        
        nuevos += 1
    
    conn.commit()
    conn.close()
    
    return {
        "success": True,
        "nuevos": nuevos,
        "duplicados": duplicados,
        "sin_fecha": sin_fecha,
        "total": len(orders)
    }


@app.get("/api/categorias")
async def get_categorias():
    """Obtiene todas las categorías."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT id, nombre, orden FROM categorias ORDER BY orden")
    categorias = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return categorias


@app.post("/api/categorias")
async def create_categoria(nombre: str = Form(...)):
    """Crea una nueva categoría."""
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT MAX(orden) FROM categorias")
        max_orden = cursor.fetchone()[0] or 0
        cursor.execute("INSERT INTO categorias (nombre, orden) VALUES (?, ?)", (nombre, max_orden + 1))
        conn.commit()
        return {"success": True, "id": cursor.lastrowid}
    except sqlite3.IntegrityError:
        raise HTTPException(400, "La categoría ya existe")
    finally:
        conn.close()


@app.get("/api/productos-sin-categoria")
async def get_productos_sin_categoria():
    """Obtiene productos que no tienen categoría asignada."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT DISTINCT lp.producto
        FROM lineas_pedido lp
        LEFT JOIN producto_categoria pc ON lp.producto = pc.producto
        WHERE pc.id IS NULL
        ORDER BY lp.producto
    ''')
    productos = [row[0] for row in cursor.fetchall()]
    conn.close()
    return productos


@app.post("/api/asignar-categoria")
async def asignar_categoria(producto: str = Form(...), categoria_id: int = Form(...)):
    """Asigna un producto a una categoría."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT OR REPLACE INTO producto_categoria (producto, categoria_id)
        VALUES (?, ?)
    ''', (producto, categoria_id))
    conn.commit()
    conn.close()
    return {"success": True}


@app.get("/api/pedidos")
async def get_pedidos(fecha: Optional[str] = None, status: Optional[str] = None):
    """Obtiene pedidos filtrados por fecha y/o estado."""
    conn = get_db()
    cursor = conn.cursor()
    
    query = "SELECT * FROM pedidos WHERE 1=1"
    params = []
    
    if fecha:
        query += " AND fecha_entrega = ?"
        params.append(fecha)
    
    if status:
        query += " AND status = ?"
        params.append(status)
    
    query += " ORDER BY fecha_entrega, order_number"
    
    cursor.execute(query, params)
    pedidos = [dict(row) for row in cursor.fetchall()]
    
    # Obtener líneas de cada pedido
    for pedido in pedidos:
        cursor.execute("SELECT * FROM lineas_pedido WHERE pedido_id = ?", (pedido['id'],))
        pedido['items'] = [dict(row) for row in cursor.fetchall()]
    
    conn.close()
    return pedidos


@app.get("/api/fechas-pendientes")
async def get_fechas_pendientes():
    """Obtiene las fechas que tienen pedidos pendientes."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT fecha_entrega, COUNT(*) as cantidad
        FROM pedidos 
        WHERE status = 'pendiente'
        GROUP BY fecha_entrega
        ORDER BY fecha_entrega
    ''')
    fechas = [{"fecha": row[0], "cantidad": row[1]} for row in cursor.fetchall()]
    conn.close()
    return fechas


@app.get("/api/lista-compras/{fecha}")
async def get_lista_compras(fecha: str):
    """Genera la lista de compras para una fecha específica."""
    conn = get_db()
    cursor = conn.cursor()
    
    # Obtener productos agrupados con categoría
    cursor.execute('''
        SELECT 
            lp.producto,
            SUM(lp.cantidad) as cantidad_total,
            COALESCE(c.nombre, 'Sin Categoría') as categoria,
            COALESCE(c.orden, 999) as categoria_orden
        FROM lineas_pedido lp
        JOIN pedidos p ON lp.pedido_id = p.id
        LEFT JOIN producto_categoria pc ON lp.producto = pc.producto
        LEFT JOIN categorias c ON pc.categoria_id = c.id
        WHERE p.fecha_entrega = ? AND p.status = 'pendiente'
        GROUP BY lp.producto
        ORDER BY categoria_orden, c.nombre, lp.producto
    ''', (fecha,))
    
    items = [dict(row) for row in cursor.fetchall()]
    conn.close()
    
    # Agrupar por categoría
    por_categoria = {}
    for item in items:
        cat = item['categoria']
        if cat not in por_categoria:
            por_categoria[cat] = []
        por_categoria[cat].append({
            'producto': item['producto'],
            'cantidad': item['cantidad_total']
        })
    
    return por_categoria


@app.get("/descargar/lista-compras/{fecha}")
async def descargar_lista_compras(fecha: str):
    """Genera y descarga Excel con lista de compras."""
    lista = await get_lista_compras(fecha)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Lista de Compras"
    
    # Estilos
    header_font = Font(bold=True, size=14)
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF", size=12)
    cat_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    cat_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws.merge_cells('A1:C1')
    ws['A1'] = f"Lista de Compras - {fecha}"
    ws['A1'].font = header_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Encabezados
    ws['A3'] = "Producto"
    ws['B3'] = "Cantidad"
    ws['C3'] = "✓"
    for col in ['A', 'B', 'C']:
        ws[f'{col}3'].font = header_font_white
        ws[f'{col}3'].fill = header_fill
        ws[f'{col}3'].border = border
        ws[f'{col}3'].alignment = Alignment(horizontal='center')
    
    row = 4
    for categoria, productos in lista.items():
        # Fila de categoría
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'] = categoria
        ws[f'A{row}'].font = cat_font
        ws[f'A{row}'].fill = cat_fill
        ws[f'A{row}'].border = border
        row += 1
        
        # Productos
        for prod in productos:
            ws[f'A{row}'] = prod['producto']
            ws[f'B{row}'] = prod['cantidad']
            ws[f'C{row}'] = "☐"
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].border = border
            ws[f'B{row}'].alignment = Alignment(horizontal='center')
            ws[f'C{row}'].alignment = Alignment(horizontal='center')
            row += 1
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 8
    
    # Guardar
    filename = f"lista_compras_{fecha}.xlsx"
    filepath = OUTPUT_DIR / filename
    wb.save(filepath)
    
    return FileResponse(
        filepath,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.get("/descargar/pedidos-armado/{fecha}")
async def descargar_pedidos_armado(fecha: str):
    """Genera y descarga Excel con pedidos para armar."""
    pedidos = await get_pedidos(fecha=fecha, status='pendiente')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Pedidos para Armar"
    
    # Estilos
    header_font = Font(bold=True, size=14)
    header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF", size=11)
    order_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    order_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws.merge_cells('A1:D1')
    ws['A1'] = f"Pedidos para Armar - {fecha}"
    ws['A1'].font = header_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws['A2'] = f"Total: {len(pedidos)} pedidos"
    ws['A2'].font = Font(italic=True)
    
    row = 4
    for pedido in pedidos:
        # Encabezado del pedido
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = f"{pedido['order_number']} | {pedido['nombre_cliente']} | {pedido['comuna']}"
        ws[f'A{row}'].font = order_font
        ws[f'A{row}'].fill = order_fill
        ws[f'A{row}'].border = border
        row += 1
        
        # Dirección
        if pedido['direccion']:
            ws[f'A{row}'] = f"📍 {pedido['direccion']}"
            row += 1
        
        # Encabezados de productos
        ws[f'A{row}'] = "Producto"
        ws[f'B{row}'] = "Cant."
        ws[f'C{row}'] = "✓"
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = header_font_white
            ws[f'{col}{row}'].fill = header_fill
            ws[f'{col}{row}'].border = border
        row += 1
        
        # Productos del pedido
        for item in pedido['items']:
            ws[f'A{row}'] = item['producto']
            ws[f'B{row}'] = item['cantidad']
            ws[f'C{row}'] = "☐"
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].border = border
            ws[f'B{row}'].alignment = Alignment(horizontal='center')
            ws[f'C{row}'].alignment = Alignment(horizontal='center')
            row += 1
        
        row += 1  # Espacio entre pedidos
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 6
    ws.column_dimensions['D'].width = 20
    
    # Guardar
    filename = f"pedidos_armado_{fecha}.xlsx"
    filepath = OUTPUT_DIR / filename
    wb.save(filepath)
    
    return FileResponse(
        filepath,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.post("/api/pedidos/{pedido_id}/completar")
async def completar_pedido(pedido_id: int):
    """Marca un pedido como completado."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("UPDATE pedidos SET status = 'completado' WHERE id = ?", (pedido_id,))
    conn.commit()
    conn.close()
    return {"success": True}


@app.delete("/api/pedidos/{pedido_id}")
async def eliminar_pedido(pedido_id: int):
    """Elimina un pedido."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM lineas_pedido WHERE pedido_id = ?", (pedido_id,))
    cursor.execute("DELETE FROM pedidos WHERE id = ?", (pedido_id,))
    conn.commit()
    conn.close()
    return {"success": True}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
